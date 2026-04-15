"""
💎 داشبورد مشروع الجوهرة - لوحة متابعة شبكات توزيع المياه
Agreement: 611250057 | الحوية - الطائف
"""

import streamlit as st
import pandas as pd
import openpyxl
import folium
from streamlit_folium import st_folium
import plotly.express as px
import plotly.graph_objects as go
import requests, zipfile, io
from datetime import datetime, timedelta
from lxml import etree

# ═══════════════════════════════════════════════════════════════
#  CONFIG
# ═══════════════════════════════════════════════════════════════
EXCEL_URL = (
    "https://privatedeals-my.sharepoint.com/:x:/g/personal/"
    "haytham_privatedeals_onmicrosoft_com/"
    "IQB3NoyTcTL5QpOXy66FaKXqAY2c0vR6KA3qIDJeUDQvmxo"
    "?e=TBvcfr&download=1"
)
KMZ_URL = (
    "https://privatedeals-my.sharepoint.com/:u:/g/personal/"
    "haytham_privatedeals_onmicrosoft_com/"
    "IQCiy9gIIcP6TYJfT5Z3HvXuAX4VTAGjU9daHJnIxAgsAw4"
    "?e=BeUHJi&download=1"
)
REFRESH_MIN = 5

# ═══════════════════════════════════════════════════════════════
#  PAGE
# ═══════════════════════════════════════════════════════════════
st.set_page_config(page_title="مشروع الجوهرة 💎", page_icon="💎",
                   layout="wide", initial_sidebar_state="expanded")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Cairo:wght@400;600;700;900&display=swap');
html,body,[class*="css"]{font-family:'Cairo',sans-serif!important;direction:rtl}
.main{background:#0a0e1a}
.kpi{background:linear-gradient(135deg,#12182b,#1a2340);border:1px solid #c9a84c55;
     border-radius:14px;padding:18px 14px;text-align:center;color:#fff;
     box-shadow:0 4px 18px rgba(201,168,76,.12);height:100%}
.kpi-lbl{font-size:12px;color:#c9a84c;margin-bottom:5px}
.kpi-val{font-size:26px;font-weight:900}
.kpi-sub{font-size:11px;color:#aaa;margin-top:4px}
.alert-g{background:#122b1a;border:1px solid #2ecc7144;border-radius:10px;padding:12px 16px;color:#2ecc71}
.alert-a{background:#2b2212;border:1px solid #e67e2244;border-radius:10px;padding:12px 16px;color:#e67e22}
.alert-r{background:#2b1212;border:1px solid #c0392b44;border-radius:10px;padding:12px 16px;color:#e74c3c}
.sec{font-size:16px;font-weight:700;color:#c9a84c;padding:6px 0;
     border-bottom:1px solid #c9a84c33;margin:18px 0 12px}
div[data-testid="stSidebar"]{background:#080c18;border-left:1px solid #c9a84c22}
.stTabs [data-baseweb="tab-list"]{gap:4px}
.stTabs [data-baseweb="tab"]{background:#12182b;border-radius:8px 8px 0 0;
  color:#c9a84c;border:1px solid #c9a84c33;padding:8px 16px;font-family:'Cairo',sans-serif}
.stTabs [aria-selected="true"]{background:#c9a84c!important;color:#0a0e1a!important;font-weight:700}
</style>
""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════
#  HELPERS
# ═══════════════════════════════════════════════════════════════
def excel_serial(n):
    try: return datetime(1899,12,30)+timedelta(days=int(n))
    except: return None

def pct(v):   return f"{v*100:.1f}%" if v else "—"
def num(v,d=0): 
    try: return f"{v:,.{d}f}"
    except: return str(v)

def kpi(label, value, sub="", color="#c9a84c"):
    st.markdown(f"""<div class="kpi">
      <div class="kpi-lbl">{label}</div>
      <div class="kpi-val" style="color:{color}">{value}</div>
      <div class="kpi-sub">{sub}</div></div>""", unsafe_allow_html=True)

DARK = dict(template="plotly_dark", paper_bgcolor="rgba(0,0,0,0)",
            plot_bgcolor="rgba(18,24,43,.6)", font=dict(family="Cairo"))

# ═══════════════════════════════════════════════════════════════
#  LOADERS
# ═══════════════════════════════════════════════════════════════
@st.cache_data(ttl=REFRESH_MIN*60, show_spinner=False)
def download_excel_bytes(url):
    """Download Excel bytes only — bytes are serializable for st.cache_data"""
    hdrs = {"User-Agent":"Mozilla/5.0"}
    try:
        r = requests.get(url, headers=hdrs, timeout=40, allow_redirects=True)
        r.raise_for_status()
        if len(r.content) < 1000:
            return None, f"رد غير متوقع ({len(r.content)} bytes) — تأكد من رابط المشاركة"
        return r.content, None
    except Exception as e:
        return None, str(e)

def parse_wb(b):
    """Parse workbook from bytes — not cached (openpyxl object not serializable)"""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(b), data_only=True, read_only=True)
        return wb, None
    except Exception as e:
        return None, str(e)

@st.cache_data(ttl=REFRESH_MIN*60, show_spinner=False)
def load_kmz(url):
    hdrs = {"User-Agent":"Mozilla/5.0"}
    try:
        r = requests.get(url, headers=hdrs, timeout=40, allow_redirects=True)
        r.raise_for_status()
        with zipfile.ZipFile(io.BytesIO(r.content)) as z:
            kf = [f for f in z.namelist() if f.lower().endswith(".kml")]
            if not kf: return [], "لا يوجد KML داخل KMZ"
            return parse_kml(z.read(kf[0])), None
    except Exception as e:
        return [], str(e)

def parse_kml(kml):
    root = etree.fromstring(kml)
    rns = root.nsmap.get(None) or "http://www.opengis.net/kml/2.2"
    ns = f"{{{rns}}}"
    def txt(el,t): f=el.find(f".//{ns}{t}"); return f.text.strip() if f is not None and f.text else ""
    items=[]
    for pm in root.iter(f"{ns}Placemark"):
        n,d = txt(pm,"name"), txt(pm,"description")
        pt = pm.find(f".//{ns}Point/{ns}coordinates")
        if pt is not None and pt.text:
            p=pt.text.strip().split(",")
            try: items.append({"name":n,"desc":d,"lat":float(p[1]),"lon":float(p[0]),"type":"point"})
            except: pass
        poly = pm.find(f".//{ns}Polygon/{ns}outerBoundaryIs/{ns}LinearRing/{ns}coordinates")
        if poly is not None and poly.text:
            cs=[[float(t.split(",")[1]),float(t.split(",")[0])] for t in poly.text.strip().split() if len(t.split(","))>=2]
            if cs: items.append({"name":n,"desc":d,"lat":sum(c[0] for c in cs)/len(cs),"lon":sum(c[1] for c in cs)/len(cs),"coords":cs,"type":"polygon"})
        line = pm.find(f".//{ns}LineString/{ns}coordinates")
        if line is not None and line.text:
            cs=[[float(t.split(",")[1]),float(t.split(",")[0])] for t in line.text.strip().split() if len(t.split(","))>=2]
            if cs: items.append({"name":n,"desc":d,"lat":sum(c[0] for c in cs)/len(cs),"lon":sum(c[1] for c in cs)/len(cs),"coords":cs,"type":"line"})
    return items

# ═══════════════════════════════════════════════════════════════
#  PARSERS
# ═══════════════════════════════════════════════════════════════
def parse_kpis(wb):
    KEYS={"ActualProgress","ExecutedLength","IssuedPermits","ExecutedValue",
          "ProjectValue","TotalLines","OpenPermits","ActiveLines",
          "CompletedLines","NotStartedLines","PlannedProgress","TotalPermitLength",
          "ExpiringSoon","Overdue"}
    kpis={}
    ws=wb["Geo Dashboard"]
    for row in ws.iter_rows(values_only=True):
        for i,v in enumerate(row):
            if v in KEYS and i+1<len(row) and row[i+1] is not None:
                kpis[v]=row[i+1]
    return kpis

def parse_scurve(wb):
    ws=wb["Dashboard"]; rows=[]; in_sc=False
    for row in ws.iter_rows(values_only=True):
        vals=[v for v in row if v is not None]
        if "Month" in [str(v) for v in vals] and any("Planned" in str(v) for v in vals):
            in_sc=True; continue
        if in_sc:
            nums=[v for v in row if isinstance(v,(int,float))]
            if len(nums)>=2 and nums[0]>40000:
                dt=excel_serial(nums[0])
                if dt: rows.append({"date":dt,"planned":nums[1],"actual":nums[2] if len(nums)>2 else None})
            elif in_sc and len(rows)>5 and not nums: break
    return pd.DataFrame(rows) if rows else pd.DataFrame()

def parse_zone_exec(wb):
    ws=wb["Geo Dashboard"]; zones={}
    for row in ws.iter_rows(values_only=True):
        rl=list(row)
        for i,v in enumerate(rl):
            if str(v).upper().startswith("ZONE") and i+1<len(rl) and isinstance(rl[i+1],(int,float)):
                zones[str(v)]=rl[i+1]
    return zones

def parse_zone_totals(wb):
    """Read total network length per zone from Dashboard sheet Zone Summary table"""
    ws = wb["Dashboard"]; totals = {}
    found_header = False; zone_col = 1; len_col = 3
    for row in ws.iter_rows(values_only=True):
        vals = list(row)
        if not found_header:
            if "Zone" in vals and "Length (m)" in vals:
                zone_col = vals.index("Zone")
                len_col  = vals.index("Length (m)")
                found_header = True
            continue
        if found_header and len(vals) > len_col and vals[zone_col]:
            name = str(vals[zone_col]).strip().upper()
            if name.startswith("ZONE") and isinstance(vals[len_col], (int,float)):
                totals[name] = vals[len_col]
            elif name == "TOTAL":
                break
    return totals

def parse_permits(wb):
    ws=wb["all project"]; rows=[]; hdr=None
    for row in ws.iter_rows(values_only=True):
        vals=list(row)
        if hdr is None:
            if "المنطقة" in vals:
                hdr=[str(v) if v else f"c{i}" for i,v in enumerate(vals)]; continue
        if hdr and any(v is not None for v in vals):
            rows.append(dict(zip(hdr,vals)))
    return pd.DataFrame(rows) if rows else pd.DataFrame()

def parse_diameter(wb):
    ws=wb["Geo Dashboard"]; d={}
    for row in ws.iter_rows(values_only=True):
        rl=list(row)
        for i,v in enumerate(rl):
            if v in {110,160,200,250} and i+1<len(rl) and isinstance(rl[i+1],float) and rl[i+1]>0:
                d[f"Ø {v} mm"]=rl[i+1]
    return d

def parse_permit_detail(wb):
    ws=wb["Permit Status"]; rows=[]; hdr=None
    for row in ws.iter_rows(values_only=True):
        vals=list(row)
        if hdr is None:
            if "التسلسل" in vals:
                hdr=[str(v) if v else f"c{i}" for i,v in enumerate(vals)]; continue
        if hdr and isinstance(vals[0],(int,float)):
            rows.append(dict(zip(hdr,vals)))
    return pd.DataFrame(rows) if rows else pd.DataFrame()

def build_map(items):
    if not items: return None
    lats=[p["lat"] for p in items]; lons=[p["lon"] for p in items]
    m=folium.Map([sum(lats)/len(lats),sum(lons)/len(lons)],zoom_start=14,tiles="CartoDB dark_matter")
    folium.TileLayer("https://mt1.google.com/vt/lyrs=s&x={x}&y={y}&z={z}",attr="Google",name="صور جوية").add_to(m)
    folium.TileLayer("CartoDB dark_matter",name="خريطة مظلمة").add_to(m)
    C={"point":"#c9a84c","polygon":"#3498db","line":"#e74c3c"}
    for p in items:
        pop=folium.Popup(f"<div dir='rtl' style='font-family:Cairo,sans-serif'><b style='color:#c9a84c'>{p['name']}</b><br>{p.get('desc','')[:150]}</div>",max_width=280)
        c=C.get(p["type"],"#c9a84c")
        if p["type"]=="point":
            folium.CircleMarker([p["lat"],p["lon"]],radius=7,color=c,fill=True,fill_color=c,fill_opacity=.8,popup=pop,tooltip=p["name"]).add_to(m)
        elif p["type"]=="polygon":
            folium.Polygon(p["coords"],color=c,fill=True,fill_color=c,fill_opacity=.2,weight=2,popup=pop,tooltip=p["name"]).add_to(m)
        elif p["type"]=="line":
            folium.PolyLine(p["coords"],color=c,weight=3,popup=pop,tooltip=p["name"]).add_to(m)
    folium.LayerControl().add_to(m)
    return m

# ═══════════════════════════════════════════════════════════════
#  SIDEBAR
# ═══════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("""<div style="text-align:center;padding:10px 0">
      <div style="font-size:28px">💎</div>
      <div style="font-size:17px;font-weight:700;color:#c9a84c">مشروع الجوهرة</div>
      <div style="font-size:11px;color:#888">شبكات المياه · الحوية - الطائف</div></div>""",
                unsafe_allow_html=True)
    st.divider()
    uploaded = st.file_uploader("📁 رفع Excel محلي", type=["xlsx","xlsm"],
                                 label_visibility="collapsed")
    if st.button("🔄 تحديث من SharePoint", use_container_width=True):
        st.cache_data.clear(); st.rerun()
    st.caption(f"⏱ تحديث تلقائي كل {REFRESH_MIN} دقائق")
    st.caption(f"🕐 {datetime.now().strftime('%H:%M:%S')}")
    st.divider()
    st.caption("💡 عند حفظ الملف في OneDrive\nيتحدث الداشبورد تلقائياً")

# ═══════════════════════════════════════════════════════════════
#  LOAD
# ═══════════════════════════════════════════════════════════════
with st.spinner("جارٍ تحميل البيانات..."):
    if uploaded:
        b = uploaded.read()
        wb, wb_err = parse_wb(b)
        src = "📁 ملف محلي"
    else:
        b, wb_err = download_excel_bytes(EXCEL_URL)
        if b is not None:
            wb, wb_err = parse_wb(b)
        else:
            wb = None
        src = "☁️ SharePoint Live"
    placemarks, kmz_err = load_kmz(KMZ_URL)

if wb is None:
    st.error(f"⚠️ تعذّر تحميل Excel: {wb_err}")
    st.info("**الحل:** استخدم زر 📎 في الشريط الجانبي لرفع الملف مباشرة.")
    st.stop()

K    = parse_kpis(wb)
SC   = parse_scurve(wb)
ZE   = parse_zone_exec(wb)
ZT   = parse_zone_totals(wb)
PM   = parse_permits(wb)
DIAM = parse_diameter(wb)
PD   = parse_permit_detail(wb)

# ═══════════════════════════════════════════════════════════════
#  HEADER
# ═══════════════════════════════════════════════════════════════
h1,h2,h3 = st.columns([5,3,2])
with h1:
    st.markdown("## 💎 داشبورد مشروع الجوهرة")
    st.caption("شبكات توزيع المياه  ●  الحوية - الطائف  ●  اتفاقية: 611250057")
with h2:
    st.markdown("""<div style='background:#12182b;border-radius:10px;padding:10px 14px;
    font-size:12px;color:#aaa;margin-top:8px'>
    📅 البداية: <b style='color:#c9a84c'>13/08/2025</b> &nbsp;
    📅 النهاية: <b style='color:#c9a84c'>12/08/2027</b></div>""", unsafe_allow_html=True)
with h3:
    st.markdown(f"""<div style='background:#12182b;border-radius:10px;padding:10px 14px;
    font-size:12px;margin-top:8px;text-align:center'>
    {'🟢' if wb else '🔴'} {src} &nbsp; {'🟢' if placemarks else '🔴'} KMZ</div>""",
    unsafe_allow_html=True)
st.divider()

# ═══════════════════════════════════════════════════════════════
#  TABS
# ═══════════════════════════════════════════════════════════════
T1,T2,T3,T4,T5 = st.tabs(["📊 ملخص المشروع","🏗️ المناطق","📋 التصاريح","🗺️ الخريطة","📈 تحليلات"])

# ──────── TAB 1 ────────
with T1:
    actual  = K.get("ActualProgress", 0) or 0
    planned = K.get("PlannedProgress", 0) or 0
    gap     = actual - planned

    c1,c2,c3,c4,c5,c6 = st.columns(6)
    with c1: kpi("الإنجاز الفعلي", pct(actual), f"الهدف: {pct(planned)}",
                 "#2ecc71" if actual>=planned else "#e74c3c")
    with c2: kpi("الفجوة عن الخطة", pct(abs(gap)),
                 "✅ متقدم" if gap>=0 else "❌ متأخر",
                 "#2ecc71" if gap>=0 else "#e74c3c")
    with c3:
        el=K.get("ExecutedLength",0) or 0; tpl=K.get("TotalPermitLength",1) or 1
        kpi("الطول المنفذ", f"{num(el)} م", f"من {num(tpl)} م", "#3498db")
    with c4: kpi("قيمة المشروع", f"{num(K.get('ProjectValue',0))} ﷼", "", "#9b59b6")
    with c5: kpi("التصاريح الصادرة", str(int(K.get("IssuedPermits",0) or 0)),
                 f"مفتوحة: {int(K.get('OpenPermits',0) or 0)}", "#e67e22")
    with c6:
        comp=int(K.get("CompletedLines",0) or 0); tot=int(K.get("TotalLines",0) or 0)
        kpi("الخطوط المنجزة", f"{comp}/{tot}",
            f"نشطة: {int(K.get('ActiveLines',0) or 0)}", "#c9a84c")

    st.markdown("")
    a1,a2,a3,a4 = st.columns(4)
    with a1: st.markdown(f'<div class="alert-g">🟢 <b>مكتملة</b><br><span style="font-size:24px;font-weight:900">{int(K.get("CompletedLines",0) or 0)}</span> خط</div>',unsafe_allow_html=True)
    with a2: st.markdown(f'<div class="alert-a">🟡 <b>نشطة</b><br><span style="font-size:24px;font-weight:900">{int(K.get("ActiveLines",0) or 0)}</span> خط</div>',unsafe_allow_html=True)
    with a3: st.markdown(f'<div class="alert-a">⏳ <b>تنتهي &lt;14 يوم</b><br><span style="font-size:24px;font-weight:900">{int(K.get("ExpiringSoon",0) or 0)}</span> تصريح</div>',unsafe_allow_html=True)
    with a4: st.markdown(f'<div class="alert-r">🔴 <b>متأخرة</b><br><span style="font-size:24px;font-weight:900">{int(K.get("Overdue",0) or 0)}</span> تصريح</div>',unsafe_allow_html=True)

    st.markdown("")
    cs, cd = st.columns([2,1])

    with cs:
        st.markdown('<div class="sec">📈 منحنى S — الخطة مقابل الفعلي</div>', unsafe_allow_html=True)
        if not SC.empty:
            today = datetime.now()
            sp = SC[SC["date"] <= today+timedelta(days=60)].copy()
            fig = go.Figure()
            fig.add_trace(go.Scatter(x=sp["date"], y=sp["planned"]*100, name="الخطة",
                line=dict(color="#3498db",width=2,dash="dash"),
                fill="tozeroy", fillcolor="rgba(52,152,219,.08)"))
            fig.add_trace(go.Scatter(x=sp["date"],
                y=[v*100 if v else None for v in sp["actual"]], name="الفعلي",
                line=dict(color="#c9a84c",width=3),
                fill="tozeroy", fillcolor="rgba(201,168,76,.12)"))
            fig.add_trace(go.Scatter(
                x=[today, today], y=[0, 105],
                mode="lines",
                line=dict(color="#e74c3c", dash="dot", width=1.5),
                name="اليوم", showlegend=False
            ))
            fig.update_layout(**DARK, height=320, margin=dict(l=0,r=0,t=10,b=0),
                yaxis=dict(title="%",range=[0,105]),
                legend=dict(orientation="h",y=1.05))
            st.plotly_chart(fig, use_container_width=True)

    with cd:
        st.markdown('<div class="sec">🔵 توزيع الأقطار</div>', unsafe_allow_html=True)
        if DIAM:
            fig_d = px.pie(names=list(DIAM.keys()), values=list(DIAM.values()),
                           color_discrete_sequence=["#c9a84c","#3498db","#2ecc71","#e74c3c"],
                           hole=0.5)
            fig_d.update_layout(**DARK, height=320, margin=dict(l=0,r=0,t=0,b=0))
            fig_d.update_traces(textinfo="percent+label", textfont=dict(family="Cairo",size=12))
            st.plotly_chart(fig_d, use_container_width=True)

# ──────── TAB 2 ────────
with T2:
    st.markdown('<div class="sec">📊 الطول المنفذ مقابل الطول الكلي لكل منطقة</div>', unsafe_allow_html=True)
    if ZE:
        # Build dataframe with executed + total
        zones_list = sorted(ZE.keys())
        zdf = pd.DataFrame({
            "المنطقة": zones_list,
            "الطول المنفذ (م)": [ZE.get(z, 0) for z in zones_list],
            "الطول الكلي (م)":  [ZT.get(z, 0) for z in zones_list],
        })
        zdf["نسبة الإنجاز"] = zdf.apply(
            lambda r: f"{r['الطول المنفذ (م)']/r['الطول الكلي (م)']*100:.1f}%"
            if r["الطول الكلي (م)"] > 0 else "—", axis=1)

        # Grouped bar chart
        fig_z = go.Figure()
        fig_z.add_trace(go.Bar(
            name="الطول الكلي",
            x=zdf["المنطقة"], y=zdf["الطول الكلي (م)"],
            marker_color="#2c4a6e",
            text=zdf["الطول الكلي (م)"].apply(lambda v: f"{v:,.0f}"),
            textposition="outside", textfont=dict(family="Cairo", size=10),
        ))
        fig_z.add_trace(go.Bar(
            name="الطول المنفذ",
            x=zdf["المنطقة"], y=zdf["الطول المنفذ (م)"],
            marker_color="#c9a84c",
            text=zdf["الطول المنفذ (م)"].apply(lambda v: f"{v:,.0f}" if v>0 else ""),
            textposition="outside", textfont=dict(family="Cairo", size=10),
        ))
        fig_z.update_layout(
            **DARK, height=400, barmode="group",
            margin=dict(l=0,r=0,t=10,b=0),
            legend=dict(orientation="h", y=1.05),
            xaxis=dict(title=""), yaxis=dict(title="متر"),
        )
        st.plotly_chart(fig_z, use_container_width=True)

        # Table with totals
        st.dataframe(
            zdf[["المنطقة","الطول الكلي (م)","الطول المنفذ (م)","نسبة الإنجاز"]],
            use_container_width=True, hide_index=True)

    st.markdown('<div class="sec">🔍 تفاصيل منطقة</div>', unsafe_allow_html=True)
    zone_sheets = [s for s in wb.sheetnames if s.upper().startswith("ZONE")]
    if zone_sheets:
        sel = st.selectbox("اختر منطقة", zone_sheets, label_visibility="collapsed")
        try:
            ws_z = wb[sel]; rows_z=[]; hdr_z=None
            for row in ws_z.iter_rows(values_only=True):
                vals=list(row)
                if hdr_z is None:
                    if "التسلسل" in vals:
                        hdr_z=[str(v) if v else f"c{i}" for i,v in enumerate(vals)]; continue
                if hdr_z and isinstance(vals[0],(int,float)):
                    rows_z.append(dict(zip(hdr_z,vals)))
            if rows_z:
                df_z = pd.DataFrame(rows_z)
                show = [c for c in ["التسلسل","رقم الخط","القطر","نوع الشارع",
                                    "طول التركيب","طول الاسفلت"] if c in df_z.columns]
                st.dataframe(df_z[show].head(60), use_container_width=True,
                             hide_index=True, height=300)
                ca,cb = st.columns(2)
                if "نوع الشارع" in df_z.columns:
                    vc=df_z["نوع الشارع"].value_counts()
                    fig_st=px.pie(names=vc.index,values=vc.values,title="نوع الشارع",
                                  color_discrete_sequence=["#c9a84c","#3498db"],hole=.4)
                    fig_st.update_layout(**DARK,height=260,margin=dict(l=0,r=0,t=40,b=0))
                    ca.plotly_chart(fig_st, use_container_width=True)
                if "القطر" in df_z.columns:
                    vc2=df_z["القطر"].value_counts().sort_index()
                    fig_dz=px.bar(x=[f"Ø{v}" for v in vc2.index],y=vc2.values,
                                  title="الأقطار",color_discrete_sequence=["#c9a84c"])
                    fig_dz.update_layout(**DARK,height=260,margin=dict(l=0,r=0,t=40,b=0))
                    cb.plotly_chart(fig_dz, use_container_width=True)
        except Exception as e:
            st.warning(str(e))

# ──────── TAB 3 ────────
with T3:
    st.markdown('<div class="sec">📋 قائمة التصاريح</div>', unsafe_allow_html=True)
    if not PM.empty:
        sc_col = next((c for c in PM.columns if "حالة التصريح" in c), None)
        if sc_col:
            vc = PM[sc_col].value_counts()
            CLR = {"إخلاء موقع":"#2ecc71","تم الانتهاء":"#2ecc71",
                   "ساري":"#e67e22","صدر تصريح":"#3498db","لم يبدأ":"#7f8c8d"}
            cols_a = st.columns(min(len(vc),4))
            for i,(s,c) in enumerate(vc.items()):
                if i<4:
                    with cols_a[i]: kpi(str(s),str(c),"خط",CLR.get(str(s),"#c9a84c"))
            st.markdown("")

        srch = st.text_input("🔍 بحث", placeholder="منطقة / رقم تصريح / خط...")
        df_show = PM.copy()
        if srch:
            mask = df_show.apply(lambda col: col.astype(str).str.contains(srch,case=False,na=False)).any(axis=1)
            df_show = df_show[mask]

        KEY = ["م","المنطقة","رقم التصريح ","الخط","القطر بالمم",
               "الطول بالمتر بالمخطط","الطول في التصريح",
               "حالة التصريح","حالة العمل بالتصريح "]
        show = [c for c in KEY if c in df_show.columns]
        st.dataframe(df_show[show] if show else df_show,
                     use_container_width=True, hide_index=True, height=450)
        st.download_button("⬇️ تحميل CSV",
            df_show.to_csv(index=False,encoding="utf-8-sig").encode("utf-8-sig"),
            "permits.csv","text/csv")
    else:
        st.info("لا توجد بيانات تصاريح")

# ──────── TAB 4 ────────
with T4:
    st.markdown('<div class="sec">🗺️ خريطة المشروع</div>', unsafe_allow_html=True)
    GMAP = "https://www.google.com/maps/d/embed?mid=1ei9SCSRMX8W6sjuzLmaIbXUYBSy3JsM"
    msrc = st.radio("مصدر الخريطة",["🗺️ Google Maps","📍 KMZ المرفوع"],horizontal=True)
    if "Google" in msrc:
        st.markdown(f'<iframe src="{GMAP}" width="100%" height="560" style="border:0;border-radius:12px" allowfullscreen></iframe>',
                    unsafe_allow_html=True)
        st.caption("📌 من Google My Maps — يتحدث عند تعديل الخريطة مباشرة")
    else:
        if kmz_err: st.warning(f"⚠️ {kmz_err}")
        if placemarks:
            c1,c2,c3=st.columns(3)
            c1.metric("📍 نقاط",sum(1 for p in placemarks if p["type"]=="point"))
            c2.metric("⬟ مضلعات",sum(1 for p in placemarks if p["type"]=="polygon"))
            c3.metric("〰 خطوط",sum(1 for p in placemarks if p["type"]=="line"))
            m=build_map(placemarks)
            if m: st_folium(m,width="100%",height=520,returned_objects=[])
        else:
            st.info("KMZ غير متاح حالياً")

# ──────── TAB 5 ────────
with T5:
    st.markdown('<div class="sec">📊 تحليلات متقدمة</div>', unsafe_allow_html=True)
    ca,cb = st.columns(2)
    with ca:
        el=K.get("ExecutedLength",0) or 0; tpl=K.get("TotalPermitLength",1) or 1
        fig_er=go.Figure(go.Bar(
            x=["المنفذ","المتبقي"],y=[el,max(tpl-el,0)],
            marker_color=["#c9a84c","#2c3e50"],
            text=[f"{el:,.0f} م",f"{max(tpl-el,0):,.0f} م"],
            textposition="outside", textfont=dict(family="Cairo")))
        fig_er.update_layout(**DARK,title="المنفذ مقابل المتبقي (م)",
                             height=300,margin=dict(l=0,r=0,t=40,b=0))
        st.plotly_chart(fig_er, use_container_width=True)
    with cb:
        ev=K.get("ExecutedValue",0) or 0; pv=K.get("ProjectValue",1) or 1
        fig_v=go.Figure(go.Pie(labels=["منفذ","متبقي"],values=[ev,max(pv-ev,0)],
            hole=.55,marker_colors=["#c9a84c","#2c3e50"],
            textinfo="percent+label",textfont=dict(family="Cairo",size=13)))
        fig_v.update_layout(**DARK,title=f"القيمة المنفذة: {ev:,.0f} ﷼",
                            height=300,margin=dict(l=0,r=0,t=40,b=0))
        st.plotly_chart(fig_v, use_container_width=True)

    if not PD.empty:
        st.markdown('<div class="sec">📊 نسبة إنجاز كل خط</div>', unsafe_allow_html=True)
        pc="اجمالي النسبة المئوية لكافة أعمال التصريح"
        zc="المنطقة"; lc="الخط"
        if pc in PD.columns and zc in PD.columns:
            pp=PD[[zc,lc,pc]].copy()
            pp[pc]=pd.to_numeric(pp[pc],errors="coerce")*100
            pp=pp.dropna(subset=[pc]).sort_values(pc)
            if not pp.empty:
                fig_pd=px.bar(pp,x=pc,y=lc,color=zc,orientation="h",
                              color_discrete_sequence=px.colors.sequential.YlOrBr,
                              height=max(400,len(pp)*18))
                fig_pd.update_layout(**DARK,margin=dict(l=0,r=0,t=10,b=0),
                                     xaxis=dict(range=[0,105],title="%"),yaxis=dict(title=""))
                st.plotly_chart(fig_pd, use_container_width=True)

st.divider()
st.caption(f"💎 مشروع الجوهرة  ●  {datetime.now().strftime('%Y-%m-%d %H:%M')}  ●  تحديث كل {REFRESH_MIN} دقائق")
